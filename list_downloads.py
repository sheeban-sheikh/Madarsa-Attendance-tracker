import os
import glob
import openpyxl

downloads_dir = r"c:\Users\Sheeban\Downloads"
pattern = os.path.join(downloads_dir, "Madarsa_Attendance_Report_*.xlsx")

files = glob.glob(pattern)
if not files:
    print("No downloaded files found.")
    exit(0)

# Sort files by name or modification time
files.sort(key=os.path.getmtime, reverse=True)

print("List of downloaded files in Downloads directory:")
for f in files:
    size = os.path.getsize(f)
    mtime = os.path.getmtime(f)
    import datetime
    mtime_str = datetime.datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
    
    data_status = "Unknown"
    try:
        wb = openpyxl.load_workbook(f)
        sheet = wb.active
        data_status = f"Valid ({sheet.max_row} rows, {sheet.max_column} cols)"
    except Exception as e:
        data_status = f"Error: {e}"
        
    print(f"  File: {os.path.basename(f)} | Size: {size} bytes | Modified: {mtime_str} | Status: {data_status}")
